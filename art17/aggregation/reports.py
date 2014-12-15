# -*- coding: utf-8 -*-
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill, Style, Font, Alignment, Border, Side
from flask import render_template, request, Response, abort
from sqlalchemy import func, and_
from functools import wraps
from sqlalchemy.orm import joinedload

from art17.aggregation import aggregation, perm_view_reports, MIMETYPE
from art17 import models, ROLE_AGGREGATED, ROLE_DRAFT, ROLE_FINAL, ROLE_MISSING
from art17.aggregation.forms import WhatForm
from art17.aggregation.utils import aggregation_missing_data_report, \
    get_checklist
from art17.auth import require
from art17.lookup import CONCLUSIONS
from art17.aggregation.export import get_tables
from art17.models import STATUS_CLOSED

PRESSURES = {
    'A': 'Agricultura',
    'B': 'Silvicultura',
    'C': 'Minerit, extractia de materiale si de productie de energie',
    'D': 'Retele de comunicatii',
    'E': 'Urbanizare, dezvoltare rezidentiala si comerciala',
    'F': 'Folosirea resurselor biologice, altele decat agricultura si silvicultura',
    'G': 'Intruziuni si dezechilibre umane',
    'H': 'Poluare',
    'I': 'Specii invazive, alte probleme ale speciilor si genele',
    'J': 'Modificari ale sistemului natural',
    'K': 'Procesele naturale biotice si abiotice (fara catastrofe)',
    'L': 'Evenimente geologice, catastrofe nturale',
    'M': 'Schimbari globale',
    'U': 'presiuni si amenintari necunoscute',
}

MEASURES = {
    '1.': 'Fara masuri',
    '2.': 'Masuri in legatura cu agricultura si habitatele deschise',
    '3.': 'Masuri legate de paduri si habitate forestiere',
    '4.': 'Masuri legate de mlastini, apa dulce si habitate de coasta',
    '5.': 'Masuri legate de habitatele marine',
    '6.': 'Masuri legate de planificarea spatiala',
    '7.': 'Masuri legate de vanatoare si pescuit si managementul speciilor',
    '8.': 'Masuri legate de ariile urbane, industriale, energie si transport',
    '9.': 'Masuri legate de utilizarea unor resurse speciale',
}

EFFECTS = [
    'broad_evaluation_maintain',
    'broad_evaluation_enhance',
    'broad_evaluation_longterm',
    'broad_evaluation_noeffect',
    'broad_evaluation_unknown',
    'broad_evaluation_notevaluated',
]

METHOD_FIELDS = {
    'habitat': [
        'range_method',
        'coverage_method',
        'coverage_trend_method',
        'structure_and_functions_method',
        'natura2000_area_method',
    ],
    'species': [
        'range_method',
        'population_method',
        'population_trend_method',
        'habitat_method',
        'natura2000_population_method',
    ],
}

METHODS = [
    ('1', 'Opinie expert'),
    ('2', 'Extrapolare'),
    ('3', 'Studiu complet'),
    ('0', 'Lipsa date'),
]

REPORT_NAMES = [
    u'#1 Număr de rapoarte încărcate în raport cu lista de verificare',
    u'#2 Situație cu înregistrările validate și cele nevalidate',
    u'#3 Numărul de rapoarte pentru habitate și specii',
    u'#4 Evaluarea globală a statutului de conservare a speciilor și habitatelor',
    u'#5 Evaluarea globală a statutului de conservare a speciilor și habitatelor pe bioregiuni',
    u'#6 Statutul de conservare a speciilor și habitatelor',
    u'#7 Modificarea statutului de conservare procentual',
    u'#8 Îmbunătățirea/deteriorarea trendului habitatelor și speciilor',
    u'#9 Evaluarea globală a statutului de conservare pe grupuri de specii și habitate procentual',
    u'#10 Motivele modificărilor valorilor parametrilor raportați între două perioade de raportare succesive',
    u'#11 Frecvența principalelor Presiuni/Amenințări',
    u'#12 Numărul evaluărilor de habitate/specii pentru care au fost raportate măsuri cu importanță mare',
    u'#13 Procentele evaluărilor de habitate/specii pentru care a fost raportat efectul unei măsuri de conservare',
    u'#14 Rapoarte de calitate a datelor pentru specii/habitate',
    u'#15 Statistici pentru specii/habitate',
]

UNKNOWN_TREND = 'x'
UNKNOWN_CONCLUSION = 'XX'
MISSING_METHOD = '0'


def get_report_name(page, dataset_id):
    dataset = models.Dataset.query.get_or_404(dataset_id)
    report_name = page or 'report'
    if dataset.year_start and dataset.year_end:
        report_name = '{}_{}-{}'.format(report_name,
                                        dataset.year_start,
                                        dataset.year_end)
    return report_name


def get_ordered_measures_codes():
    codes = MEASURES.keys()
    codes.sort()
    return codes


def get_report_data(dataset, roles=None):
    roles = roles or (ROLE_AGGREGATED, ROLE_DRAFT, ROLE_FINAL)
    species = (
        dataset.species_objs
        .join(models.DataSpeciesRegion.species)
        .options(
            joinedload(models.DataSpeciesRegion.species)
        ).filter(models.DataSpeciesRegion.cons_role.in_(roles))
    )
    habitats = (
        dataset.habitat_objs
        .join(models.DataHabitattypeRegion.habitat)
        .options(
            joinedload(models.DataHabitattypeRegion.habitat)
        ).filter(models.DataHabitattypeRegion.cons_role.in_(roles))
    )
    return species, habitats


def get_checklist_data(dataset):
    checklist = get_checklist(dataset.checklist_id)
    species = (
        models.DataSpeciesCheckList.query.filter_by(
            dataset_id=checklist.id, member_state='RO',
        )
    )
    habitats = (
        models.DataHabitatsCheckList.query.filter_by(
            dataset_id=checklist.id, member_state='RO',
        )
    )
    return species, habitats


def get_checklist_map(dataset):
    species_cl, habitats_cl = get_checklist_data(dataset)
    return _get_map(species_cl), _get_map(habitats_cl)


def _get_map(checklist_qs):
    return {
        (s.code, s.bio_region): s for s in checklist_qs
    }


def get_measures_count(data_query, attr_name):
    return (
        data_query.join(models.DataMeasures)
        .filter(
            models.DataMeasures.lu_ranking.has(name='high importance'),
            getattr(models.DataMeasures, attr_name) != None,
        ).with_entities(
            models.DataMeasures.measurecode,
            func.count(models.DataMeasures.id),
        ).group_by(models.DataMeasures.measurecode)
    ).all()


def get_measures(data_query, attr_name):
    return (
        data_query.join(models.DataMeasures)
        .filter(getattr(models.DataMeasures, attr_name) != None)
        .with_entities(
            models.DataMeasures.measurecode,
            models.DataMeasures.broad_evaluation_maintain,
            models.DataMeasures.broad_evaluation_enhance,
            models.DataMeasures.broad_evaluation_longterm,
            models.DataMeasures.broad_evaluation_noeffect,
            models.DataMeasures.broad_evaluation_unknown,
            models.DataMeasures.broad_evaluation_notevaluated,
        ))


def add_to_measures_dict(measures_dict, measures_query, category):
    for measure_code, reports_count in measures_query:
        code = measure_code[:2]
        if code == u'1':
            code = u'1.'
        measures_dict[code][category] += reports_count


def get_effects_dict(data_measures):
    effects_dict = {code: {effect: 0 for effect in EFFECTS}
                    for code in MEASURES.keys()}
    measures_dict = {code: 0 for code in MEASURES.keys()}

    for data_measure in data_measures:
        categ_code = data_measure.measurecode[:2]
        if categ_code == u'1':
            categ_code = u'1.'
        measures_dict[categ_code] += 1
        for effect in EFFECTS:
            effects_dict[categ_code][effect] += getattr(data_measure, effect)

    for measure_code, total in measures_dict.iteritems():
        for effect in EFFECTS:
            res = effects_dict[measure_code][effect] * 100 / total \
                if total else Decimal(0)
            effects_dict[measure_code][effect] = res.quantize(Decimal('1.00'))
    return effects_dict


def get_methods_quality_dict(data_query, data_class, category):
    count = Decimal(data_query.count() or 1)
    methods_quality = {
        method: {
            field: (
                data_query.filter(getattr(data_class, field) == method).count()
                * 100 / count
            ).quantize(Decimal('1.00'))
            for field in METHOD_FIELDS[category]
        }
        for method, _ in METHODS
    }
    for method, fields in methods_quality.iteritems():
        methods_quality[method]['average'] = sum(fields.values()) / len(fields)
    return methods_quality


def get_excel_document(html, filename):
    tables = get_tables(html)
    wb = Workbook()
    for _ in range(len(tables) - len(wb.worksheets)):
        wb.create_sheet()
    style = Style(
        font=Font(bold=True),
        fill=PatternFill(start_color='EEEEEE', end_color='EEEEEE',
                         fill_type='solid'),
        border=Border(left=Side(border_style='thin'),
                      right=Side(border_style='thin'),
                      top=Side(border_style='thin'),
                      bottom=Side(border_style='thin')),
        alignment=Alignment(horizontal='center', vertical='center'),
    )

    for table, sheet in zip(tables, wb.worksheets):
        sheet.title = table.title
        for row in table.rows:
            for cell in row.cells:
                sheet.merge_cells(
                    start_row=cell.row.idx,
                    start_column=cell.col_idx,
                    end_row=cell.row.idx + cell.rowspan,
                    end_column=cell.col_idx + cell.colspan)
                col_letter = get_column_letter(cell.col_idx)
                sheet_cell = sheet[col_letter + str(cell.row.idx)]
                sheet_cell.value = cell.text
                if cell.tag == 'th':
                    sheet_cell.style = style

    resp = Response(save_virtual_workbook(wb), mimetype=MIMETYPE)
    resp.headers.add('Content-Disposition',
                     'attachment; filename={}.xlsx'.format(filename))
    return resp


def download():
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            html = func(*args, **kwargs)
            if request.args.get('download'):
                page = kwargs.get('page')
                dataset_id = kwargs.get('dataset_id')
                report_name = get_report_name(page, dataset_id)
                return get_excel_document(html, report_name)
            return html

        return wrapper

    return decorator


@aggregation.app_context_processor
def inject_consts():
    groups = dict(
        models.LuGrupSpecie.query
        .with_entities(models.LuGrupSpecie.code,
                       models.LuGrupSpecie.description)
    )
    return dict(GROUPS=groups, REPORT_NAMES=REPORT_NAMES)


@aggregation.route('/raport/<int:dataset_id>')
@require(perm_view_reports())
def report(dataset_id):
    dataset = models.Dataset.query.get_or_404(dataset_id)
    return render_template('aggregation/reports/report.html',
                           dataset=dataset,
                           dataset_id=dataset.id)


def report_conservation_status(dataset_id, page):
    dataset = models.Dataset.query.get_or_404(dataset_id)
    species, habitats = get_report_data(dataset)
    species_cl, habitats_cl = get_checklist_map(dataset)
    species = list(
        species
        .join(models.DataSpeciesRegion.species)
        .join(models.DataSpecies.lu)
        .join(models.LuHdSpecies.group)
        .order_by(models.LuGrupSpecie.description, models.DataSpecies.code)
    )
    habitats = list(
        habitats
        .join(models.DataHabitattypeRegion.habitat)
        .order_by(models.DataHabitat.code)
    )
    all_species = len(species) or 1
    all_habitats = len(habitats) or 1
    stats = {
        'species': {'range': {}, 'population': {}, 'habitat': {}, 'future': {},
                    'assessment': {}},
        'habitats': {'range': {}, 'area': {}, 'structure': {}, 'future': {},
                     'assessment': {}}
    }

    def set_stat(conclusion, stats):
        conclusion = conclusion or '-'
        stats.setdefault(conclusion, 0)
        stats[conclusion] += 1

    for spec in species:
        set_stat(spec.conclusion_range, stats['species']['range'])
        set_stat(spec.conclusion_population, stats['species']['population'])
        set_stat(spec.conclusion_habitat, stats['species']['habitat'])
        set_stat(spec.conclusion_future, stats['species']['future'])
        set_stat(spec.conclusion_assessment, stats['species']['assessment'])
    for hab in habitats:
        set_stat(hab.conclusion_range, stats['habitats']['range'])
        set_stat(hab.conclusion_area, stats['habitats']['area'])
        set_stat(hab.conclusion_structure, stats['habitats']['structure'])
        set_stat(hab.conclusion_future, stats['habitats']['future'])
        set_stat(hab.conclusion_assessment, stats['habitats']['assessment'])
    for k, values in stats['species'].iteritems():
        for conclusion, v in values.iteritems():
            stats['species'][k][conclusion] = v * 100.0 / all_species
    for k, values in stats['habitats'].iteritems():
        for conclusion, v in values.iteritems():
            stats['habitats'][k][conclusion] = v * 100.0 / all_habitats

    groups = dict(
        models.LuGrupSpecie.query
        .with_entities(models.LuGrupSpecie.code,
                       models.LuGrupSpecie.description)
    )

    return render_template(
        'aggregation/reports/conservation_status.html',
        dataset=dataset, dataset_id=dataset.id,
        species=species, habitats=habitats, page=page, stats=stats,
        species_cl=species_cl, habitats_cl=habitats_cl,
        GROUPS=groups,
    )


def report_bioreg_global(dataset_id, page):
    dataset = models.Dataset.query.get_or_404(dataset_id)
    species, habitats = get_report_data(dataset)

    all_species = species.count()
    all_habitats = habitats.count()

    REGIONS = ('ALP', 'CON', 'PAN', 'STE', 'BLS', 'MBLS')
    stats = {
        'species': {r: {} for r in REGIONS},
        'habitats': {r: {} for r in REGIONS}
    }

    for spec in species:
        conclusion = spec.conclusion_assessment or 'NA'
        stats['species'].setdefault(spec.region, {})
        stats['species'][spec.region].setdefault(conclusion, 0)
        stats['species'][spec.region][conclusion] += 1

    for hab in habitats:
        conclusion = hab.conclusion_assessment or 'NA'
        stats['habitats'].setdefault(hab.region, {})
        stats['habitats'][hab.region].setdefault(conclusion, 0)
        stats['habitats'][hab.region][conclusion] += 1

    for k, values in stats['species'].iteritems():
        all_species = sum(stats['species'][k].values()) or 1
        for conclusion, v in values.iteritems():
            stats['species'][k][conclusion] = v * 100.0 / all_species
    for k, values in stats['habitats'].iteritems():
        all_habitats = sum(stats['habitats'][k].values()) or 1
        for conclusion, v in values.iteritems():
            stats['habitats'][k][conclusion] = v * 100.0 / all_habitats

    return render_template(
        'aggregation/reports/bioreg_global.html',
        dataset=dataset, dataset_id=dataset.id, regions=REGIONS,
        species=species, habitats=habitats, page=page, stats=stats)


def report_bioreg_annex(dataset_id, page):
    dataset = models.Dataset.query.get_or_404(dataset_id)
    species_cl, habitats_cl = get_checklist_data(dataset)

    if request.args.get('what'):
        what_form = WhatForm(request.args)
    else:
        what_form = WhatForm()
    if what_form.what.data == 1:
        roles = (ROLE_FINAL,)
    elif what_form.what.data == 2:
        roles = (ROLE_AGGREGATED, ROLE_DRAFT)
    else:
        roles = None
    species, habitats = get_report_data(dataset, roles=roles)

    def _get_regions(queryset):
        regions = (
            queryset
            .filter_by(member_state='RO')
            .with_entities('bio_region')
            .distinct()
        )
        return [r for r, in regions]

    species_regions = _get_regions(species_cl)
    habitats_regions = _get_regions(habitats_cl)
    regions = list(set(species_regions + habitats_regions))
    stats = {
        'species': {
            r: {annex: {'n': 0, 'p': 0} for annex in (2, 4, 5)}
            for r in regions
        },
        'habitats': {r: {'n': 0, 'p': 0} for r in regions}
    }
    species_cl = {
        (s.code, s.bio_region): s for s in species_cl
    }
    habitats_cl = {
        (h.code, h.bio_region): h for h in habitats_cl
    }

    for s in species:
        spec = species_cl.get((s.species and s.species.code, s.region), None)
        if not spec:
            continue
        if spec.has_annex(2):
            if spec.priority:
                stats['species'][spec.bio_region][2]['p'] += 1
            else:
                stats['species'][spec.bio_region][2]['n'] += 1
            for annex in (4, 5):
                if spec.has_annex(annex):
                    stats['species'][spec.bio_region][annex]['p'] += 1
        else:
            for annex in (4, 5):
                if spec.has_annex(annex):
                    stats['species'][spec.bio_region][annex]['n'] += 1
    for h in habitats:
        hab = habitats_cl.get((h.habitat and h.habitat.code, h.region), None)
        if not hab:
            continue
        if hab.priority:
            stats['habitats'][hab.bio_region]['p'] += 1
        else:
            stats['habitats'][hab.bio_region]['n'] += 1

    return render_template(
        'aggregation/reports/bioreg_annex.html',
        dataset=dataset, dataset_id=dataset.id, regions=regions,
        species=species_cl, habitats=habitats_cl, page=page,
        stats=stats,
        current_checklist=get_checklist(dataset.checklist_id),
        what_form=what_form,
    )


def report_pressures1(dataset_id, page):
    dataset = models.Dataset.query.get_or_404(dataset_id)
    species, habitats = get_report_data(dataset)

    species_pres = (
        species.join(models.DataPressuresThreats)
        .filter(models.DataPressuresThreats.type == 'p')
        .with_entities(models.DataPressuresThreats.pressure,
                       func.count(models.DataPressuresThreats.id))
        .group_by(models.DataPressuresThreats.pressure,
                  models.DataPressuresThreats.type)
    ).all()
    species_thr = (
        species.join(models.DataPressuresThreats)
        .filter(models.DataPressuresThreats.type == 't')
        .with_entities(models.DataPressuresThreats.pressure,
                       func.count(models.DataPressuresThreats.id))
        .group_by(models.DataPressuresThreats.pressure,
                  models.DataPressuresThreats.type)
    ).all()
    habitat_pres = (
        habitats.join(models.DataPressuresThreats)
        .filter(models.DataPressuresThreats.type == 'p')
        .with_entities(models.DataPressuresThreats.pressure,
                       func.count(models.DataPressuresThreats.id))
        .group_by(models.DataPressuresThreats.pressure)
    ).all()
    habitat_thr = (
        habitats.join(models.DataPressuresThreats)
        .filter(models.DataPressuresThreats.type == 't')
        .with_entities(models.DataPressuresThreats.pressure,
                       func.count(models.DataPressuresThreats.id))
        .group_by(models.DataPressuresThreats.pressure)
    ).all()

    stats = {}
    for k in PRESSURES:
        stats.setdefault(k, {})
        stats[k].setdefault('species', {})
        stats[k]['species']['pressures'] = sum(
            [val for name, val in species_pres if name and name.startswith(k)])
        stats[k]['species']['threats'] = sum(
            [val for name, val in species_thr if name and name.startswith(k)])
        stats[k].setdefault('habitats', {})
        stats[k]['habitats']['pressures'] = sum(
            [val for name, val in habitat_pres if name and name.startswith(k)])
        stats[k]['habitats']['threats'] = sum(
            [val for name, val in habitat_thr if name and name.startswith(k)])
    all_spec_pres = sum(
        [stats[k]['species']['pressures'] for k in PRESSURES]) or 1
    all_spec_thr = sum(
        [stats[k]['species']['threats'] for k in PRESSURES]) or 1
    all_hab_pres = sum(
        [stats[k]['habitats']['pressures'] for k in PRESSURES]) or 1
    all_hab_thr = sum(
        [stats[k]['habitats']['threats'] for k in PRESSURES]) or 1
    for k in PRESSURES:
        stats[k]['species']['pressures'] = stats[k]['species'][
                                               'pressures'] * 100.0 / all_spec_pres
        stats[k]['species']['threats'] = stats[k]['species'][
                                             'threats'] * 100.0 / all_spec_thr
        stats[k]['habitats']['pressures'] = stats[k]['habitats'][
                                                'pressures'] * 100.0 / all_hab_pres
        stats[k]['habitats']['threats'] = stats[k]['habitats'][
                                              'threats'] * 100.0 / all_hab_thr

    return render_template(
        'aggregation/reports/pressures1.html',
        dataset=dataset, dataset_id=dataset.id, pressures=PRESSURES,
        species=species, habitats=habitats, page=page, stats=stats)


def report_measures_high_importance(dataset_id, page):
    dataset = models.Dataset.query.get_or_404(dataset_id)
    species, habitats = get_report_data(dataset)

    species_measures_count = get_measures_count(species, 'species_id')
    habitat_measures_count = get_measures_count(habitats, 'habitat_id')
    measures_dict = {code: {'species': 0, 'habitat': 0}
                     for code in MEASURES.keys()}
    add_to_measures_dict(measures_dict, species_measures_count, 'species')
    add_to_measures_dict(measures_dict, habitat_measures_count, 'habitat')
    ordered_keys = measures_dict.keys()
    ordered_keys.sort()

    return render_template(
        'aggregation/reports/measures.html',
        dataset=dataset, count=measures_dict, names=MEASURES,
        ordered_keys=get_ordered_measures_codes(), page=page, )


def report_missing(dataset_id, page):
    dataset = models.Dataset.query.get_or_404(dataset_id)

    species = dataset.species_objs.filter_by(cons_role=ROLE_MISSING)
    habitats = dataset.habitat_objs.filter_by(cons_role=ROLE_MISSING)
    species_cl, habitats_cl = get_checklist_map(dataset)

    groups = dict(
        models.LuGrupSpecie.query
        .with_entities(models.LuGrupSpecie.code,
                       models.LuGrupSpecie.description)
    )

    return render_template(
        'aggregation/reports/missing.html', dataset=dataset,
        page=page, missing_species=species, missing_habitats=habitats,
        GROUPS=groups, species_cl=species_cl, habitats_cl=habitats_cl,
    )


def report_measures_effects(dataset_id, page):
    dataset = models.Dataset.query.get_or_404(dataset_id)
    species, habitat = get_report_data(dataset)

    species_data_measures = get_measures(species, 'species_id')
    habitat_data_measures = get_measures(habitat, 'habitat_id')

    species_effects_dict = get_effects_dict(species_data_measures)
    habitat_effects_dict = get_effects_dict(habitat_data_measures)

    return render_template(
        'aggregation/reports/measures_effects.html',
        dataset=dataset, species_count=species_effects_dict,
        habitat_count=habitat_effects_dict, names=MEASURES,
        ordered_keys=get_ordered_measures_codes(), page=page, )


def report_quality(dataset_id, page):
    dataset = models.Dataset.query.get_or_404(dataset_id)
    species, habitat = get_report_data(dataset)

    habitat_methods_quality = get_methods_quality_dict(
        habitat, models.DataHabitattypeRegion, 'habitat')
    species_methods_quality = get_methods_quality_dict(
        species, models.DataSpeciesRegion, 'species')

    habitat_unknown_data = {
        'range_surface_area': {
            'missing': habitat.filter(
                models.DataHabitattypeRegion.range_method == MISSING_METHOD).count(),
            'unknown': habitat.filter(
                models.DataHabitattypeRegion.range_surface_area == None).count(),
        },
        'range_trend': {
            'missing': habitat.filter(
                models.DataHabitattypeRegion.range_trend == None).count(),
            'unknown': habitat.filter(
                models.DataHabitattypeRegion.range_trend == UNKNOWN_TREND).count(),
        },
        'complementary_favourable_range': {
            'missing': (habitat
                        .filter(and_(
                models.DataHabitattypeRegion.complementary_favourable_range == None,
                models.DataHabitattypeRegion.complementary_favourable_range_op == None,
                models.DataHabitattypeRegion.complementary_favourable_range_unknown == None
            )
            ).count()),
            'unknown': (habitat
                        .filter(
                models.DataHabitattypeRegion.complementary_favourable_range_unknown == 1)
                        .count()),
        },
        'conclusion_range': {
            'missing': habitat.filter(
                models.DataHabitattypeRegion.conclusion_range == None).count(),
            'unknown': habitat.filter(
                models.DataHabitattypeRegion.conclusion_range == UNKNOWN_CONCLUSION,
            ).count(),
        },
        'coverage_surface_area': {
            'missing': habitat.filter(
                models.DataHabitattypeRegion.coverage_method == MISSING_METHOD,
            ).count(),
            'unknown': habitat.filter(
                models.DataHabitattypeRegion.coverage_surface_area == None
            ).count(),
        },
        'coverage_trend': {
            'missing': habitat.filter(
                models.DataHabitattypeRegion.coverage_trend == None).count(),
            'unknown': habitat.filter(
                models.DataHabitattypeRegion.coverage_trend == UNKNOWN_TREND).count(),
        },
        'complementary_favourable_area': {
            'missing': habitat.filter(and_(
                models.DataHabitattypeRegion.complementary_favourable_area == None,
                models.DataHabitattypeRegion.complementary_favourable_area_op == None,
                models.DataHabitattypeRegion.complementary_favourable_area_unknown == None)
            ).count(),
            'unknown': habitat.filter(
                models.DataHabitattypeRegion.complementary_favourable_area_unknown == 1)
                .count(),
        },
        'conclusion_area': {
            'missing': habitat.filter(
                models.DataHabitattypeRegion.conclusion_area == None).count(),
            'unknown': habitat.filter(
                models.DataHabitattypeRegion.conclusion_area == UNKNOWN_CONCLUSION,
            ).count(),
        },
        'conclusion_structure': {
            'missing': habitat.filter(
                models.DataHabitattypeRegion.conclusion_structure == None).count(),
            'unknown': habitat.filter(
                models.DataHabitattypeRegion.conclusion_structure == UNKNOWN_CONCLUSION,
            ).count(),
        },
        'conclusion_future': {
            'missing': habitat.filter(
                models.DataHabitattypeRegion.conclusion_future == None).count(),
            'unknown': habitat.filter(
                models.DataHabitattypeRegion.conclusion_future == UNKNOWN_CONCLUSION,
            ).count(),
        },
        'pressures': {
            'unknown': habitat.filter(
                models.DataHabitattypeRegion.pressures_method == None).count(),
        },
        'threats': {
            'unknown': habitat.filter(
                models.DataHabitattypeRegion.threats_method == None).count(),
        },
        'natura2000': {
            'missing': habitat.filter(and_(
                models.DataHabitattypeRegion.natura2000_area_min == None,
                models.DataHabitattypeRegion.natura2000_area_max == None,
            )).count(),
        },
        'conclusion_assessment': {
            'missing': habitat.filter(
                models.DataHabitattypeRegion.conclusion_assessment == None).count(),
            'unknown': habitat.filter(
                models.DataHabitattypeRegion.conclusion_assessment == UNKNOWN_CONCLUSION,
            ).count(),
        },
        'conclusion_assessment_trend': {
            'missing': habitat.filter(
                models.DataHabitattypeRegion.conclusion_assessment_trend == None,
            ).count(),
            'unknown': habitat.filter(
                models.DataHabitattypeRegion.conclusion_assessment_trend == UNKNOWN_TREND,
            ).count(),
        },
    }
    habitat_count = Decimal(habitat.count())
    habitat_unknown_data = {
        k: {ik: ((iv * 100 / habitat_count) if habitat_count else Decimal(0))
            .quantize(Decimal('1.00'))
            for ik, iv in v.iteritems()}
        for k, v in habitat_unknown_data.iteritems()
    }

    species_unknown_data = {
        'range_surface_area': {
            'missing': species.filter(
                models.DataSpeciesRegion.range_method == MISSING_METHOD).count(),
            'unknown': species.filter(
                models.DataSpeciesRegion.range_surface_area == None).count(),
        },
        'range_trend': {
            'missing': species.filter(
                models.DataSpeciesRegion.range_trend == None).count(),
            'unknown': species.filter(
                models.DataSpeciesRegion.range_trend == UNKNOWN_TREND).count(),
        },
        'complementary_favourable_range': {
            'missing': (species
                        .filter(and_(
                models.DataSpeciesRegion.complementary_favourable_range == None,
                models.DataSpeciesRegion.complementary_favourable_range_op == None,
                models.DataSpeciesRegion.complementary_favourable_range_unknown == None,
            )
            ).count()),
            'unknown': (species
                        .filter(
                models.DataSpeciesRegion.complementary_favourable_range_unknown == 1)
                        .count()),
        },
        'conclusion_range': {
            'missing': species.filter(
                models.DataSpeciesRegion.conclusion_range == None).count(),
            'unknown': species.filter(
                models.DataSpeciesRegion.conclusion_range == UNKNOWN_CONCLUSION,
            ).count(),
        },
        'population_size_unit': {
            'missing': species.filter(
                models.DataSpeciesRegion.population_method == MISSING_METHOD,
            ).count(),
            'unknown': species.filter(
                models.DataSpeciesRegion.population_size_unit == None).count(),
        },
        'population_trend': {
            'missing': species.filter(
                models.DataSpeciesRegion.population_trend == None).count(),
            'unknown': species.filter(
                models.DataSpeciesRegion.population_trend == UNKNOWN_TREND,
            ).count(),
        },
        'complementary_favourable_population': {
            'missing': (species.filter(and_(
                models.DataSpeciesRegion.complementary_favourable_population == None,
                models.DataSpeciesRegion.complementary_favourable_population_op == None,
                models.DataSpeciesRegion.complementary_favourable_population_unknown == None,
            )).count()),
            'unknown': (species.filter(
                models.DataSpeciesRegion.complementary_favourable_population_unknown == 1,
            ).count()),
        },
        'conclusion_population': {
            'missing': species.filter(
                models.DataSpeciesRegion.conclusion_population == None).count(),
            'unknown': species.filter(
                models.DataSpeciesRegion.conclusion_population == UNKNOWN_CONCLUSION,
            ).count(),
        },
        'habitat_surface_area': {
            'missing': species.filter(
                models.DataSpeciesRegion.habitat_method == MISSING_METHOD,
            ).count(),
            'unknown': species.filter(
                models.DataSpeciesRegion.habitat_surface_area == None).count(),
        },
        'habitat_trend': {
            'missing': species.filter(
                models.DataSpeciesRegion.habitat_trend == None).count(),
            'unknown': species.filter(
                models.DataSpeciesRegion.habitat_trend == UNKNOWN_TREND).count(),
        },
        'habitat_area_suitable': {
            'missing': species.filter(
                models.DataSpeciesRegion.habitat_area_suitable == 0).count(),
            'unknown': species.filter(
                models.DataSpeciesRegion.habitat_area_suitable == None).count(),
        },
        'conclusion_habitat': {
            'missing': species.filter(
                models.DataSpeciesRegion.conclusion_habitat == None).count(),
            'unknown': species.filter(
                models.DataSpeciesRegion.conclusion_habitat == UNKNOWN_CONCLUSION,
            ).count(),
        },
        'conclusion_future': {
            'missing': species.filter(
                models.DataSpeciesRegion.conclusion_future == None).count(),
            'unknown': species.filter(
                models.DataSpeciesRegion.conclusion_future == UNKNOWN_CONCLUSION,
            ).count(),
        },
        'pressures': {
            'unknown': species.filter(
                models.DataSpeciesRegion.pressures_method == None).count(),
        },
        'threats': {
            'unknown': species.filter(
                models.DataSpeciesRegion.threats_method == None).count(),
        },
        'natura2000': {
            'missing': species.filter(and_(
                models.DataSpeciesRegion.natura2000_population_min == None,
                models.DataSpeciesRegion.natura2000_population_max == None,
            )
            ).count(),
        },
        'conclusion_assessment': {
            'missing': species.filter(
                models.DataSpeciesRegion.conclusion_assessment == None).count(),
            'unknown': species.filter(
                models.DataSpeciesRegion.conclusion_assessment == UNKNOWN_CONCLUSION,
            ).count(),
        },
        'conclusion_assessment_trend': {
            'missing': species.filter(
                models.DataSpeciesRegion.conclusion_assessment_trend == None,
            ).count(),
            'unknown': species.filter(
                models.DataSpeciesRegion.conclusion_assessment_trend == UNKNOWN_TREND,
            ).count(),
        },
    }
    species_count = Decimal(species.count())
    species_unknown_data = {
        k: {ik: ((iv * 100 / species_count) if species_count else Decimal(0))
            .quantize(Decimal('1.00'))
            for ik, iv in v.iteritems()}
        for k, v in species_unknown_data.iteritems()
    }

    return render_template(
        'aggregation/reports/quality.html',
        dataset=dataset, fields=METHOD_FIELDS, methods=METHODS, page=page,
        habitat_methods_quality=habitat_methods_quality,
        species_methods_quality=species_methods_quality,
        habitat_unknown_data=habitat_unknown_data,
        species_unknown_data=species_unknown_data,
    )


def report_validation(dataset_id, page):
    dataset = models.Dataset.query.get_or_404(dataset_id)
    species, habitat = get_report_data(dataset)
    species_cl, habitats_cl = get_checklist_map(dataset)

    species = (
        species
        .join(models.DataSpeciesRegion.species)
        .join(models.DataSpecies.lu)
        .join(models.LuHdSpecies.group)
        .order_by(models.LuGrupSpecie.description, models.DataSpecies.code)
    )
    habitat = (
        habitat
        .join(models.DataHabitattypeRegion.habitat)
        .order_by(models.DataHabitat.code)
    )

    return render_template(
        'aggregation/reports/validation.html',
        page=page, dataset=dataset, species=species, habitats=habitat,
        species_cl=species_cl, habitats_cl=habitats_cl,
    )


def report_13(dataset_id, page):
    dataset = models.Dataset.query.get_or_404(dataset_id)

    datasets = set(
        list(models.Dataset.query.filter_by(status=STATUS_CLOSED)) + [dataset]
    )
    for ds in datasets:
        species, habitat = get_report_data(ds)

        species_count = species.count() or 1
        habitat_count = habitat.count() or 1

        species = species.with_entities(
            models.DataSpeciesRegion.conclusion_assessment,
            func.count(models.DataSpeciesRegion.id)
        ).group_by(models.DataSpeciesRegion.conclusion_assessment)
        ds.species = dict(species)
        ds.species['NA'] = ds.species.get(None, 0)
        for key, value in ds.species.iteritems():
            ds.species[key] = value * 100.0 / species_count

        habitats = habitat.with_entities(
            models.DataHabitattypeRegion.conclusion_assessment,
            func.count(models.DataHabitattypeRegion.id)
        ).group_by(models.DataHabitattypeRegion.conclusion_assessment)
        ds.habitat = dict(habitats)
        ds.habitat['NA'] = ds.habitat.get(None, 0)
        for key, value in ds.habitat.iteritems():
            ds.habitat[key] = value * 100.0 / habitat_count

    datasets = list(datasets)
    datasets.sort(key=lambda d: d.year_end, reverse=True)

    return render_template(
        'aggregation/reports/13.html',
        page=page, datasets=datasets, dataset=dataset,
    )


def report_14(dataset_id, page):
    dataset = models.Dataset.query.get_or_404(dataset_id)
    species, habitat = get_report_data(dataset)

    species_mod = species.filter(
        (models.DataSpeciesRegion.range_reasons_for_change_a == 1) |
        (models.DataSpeciesRegion.range_reasons_for_change_b == 1) |
        (models.DataSpeciesRegion.range_reasons_for_change_c == 1) |
        (models.DataSpeciesRegion.population_reasons_for_change_a == 1) |
        (models.DataSpeciesRegion.population_reasons_for_change_b == 1) |
        (models.DataSpeciesRegion.population_reasons_for_change_c == 1) |
        (models.DataSpeciesRegion.habitat_reasons_for_change_a == 1) |
        (models.DataSpeciesRegion.habitat_reasons_for_change_b == 1) |
        (models.DataSpeciesRegion.habitat_reasons_for_change_c == 1)
    ).count()
    species_real = species.filter(
        (models.DataSpeciesRegion.range_reasons_for_change_a == 1) |
        (models.DataSpeciesRegion.population_reasons_for_change_a == 1) |
        (models.DataSpeciesRegion.habitat_reasons_for_change_a == 1)
    ).count()
    species_count = species.count() or 1
    species_real *= 100.0 / species_mod
    species_mod *= 100.0 / species_count

    habitat_mod = habitat.filter(
        (models.DataHabitattypeRegion.range_reasons_for_change_a == 1) |
        (models.DataHabitattypeRegion.range_reasons_for_change_b == 1) |
        (models.DataHabitattypeRegion.range_reasons_for_change_c == 1) |
        (models.DataHabitattypeRegion.area_reasons_for_change_a == 1) |
        (models.DataHabitattypeRegion.area_reasons_for_change_b == 1) |
        (models.DataHabitattypeRegion.area_reasons_for_change_c == 1)
    ).count()
    habitat_real = habitat.filter(
        (models.DataHabitattypeRegion.range_reasons_for_change_a == 1) |
        (models.DataHabitattypeRegion.area_reasons_for_change_a == 1)
    ).count()
    habitat_count = habitat.count() or 1
    habitat_real *= 100.0 / habitat_mod
    habitat_mod *= 100.0 / habitat_count

    return render_template(
        'aggregation/reports/14.html',
        page=page, dataset=dataset, species_mod=species_mod,
        species_real=species_real, habitat_mod=habitat_mod,
        habitat_real=habitat_real,
    )


def report_15(dataset_id, page):
    dataset = models.Dataset.query.get_or_404(dataset_id)
    species, habitat = get_report_data(dataset)

    species_data = species.with_entities(
        func.concat(models.DataSpeciesRegion.conclusion_assessment,
                    models.DataSpeciesRegion.conclusion_assessment_trend),
        func.count(models.DataSpeciesRegion.id)
    ).group_by(models.DataSpeciesRegion.conclusion_assessment,
               models.DataSpeciesRegion.conclusion_assessment_trend)

    species_count = species.filter(
        models.DataSpeciesRegion.conclusion_assessment.startswith('U')
    ).count() or 1
    species_data = dict(species_data) or {}
    for key, value in species_data.iteritems():
        species_data[key] = value * 100.0 / species_count

    habitat_data = habitat.with_entities(
        func.concat(models.DataHabitattypeRegion.conclusion_assessment,
                    models.DataHabitattypeRegion.conclusion_assessment_trend),
        func.count(models.DataHabitattypeRegion.id)
    ).group_by(models.DataHabitattypeRegion.conclusion_assessment,
               models.DataHabitattypeRegion.conclusion_assessment_trend)

    habitat_count = habitat.filter(
        models.DataHabitattypeRegion.conclusion_assessment.startswith('U')
    ).count() or 1
    habitat_data = dict(habitat_data) or {}
    for key, value in habitat_data.iteritems():
        habitat_data[key] = value * 100.0 / habitat_count

    return render_template(
        'aggregation/reports/15.html',
        page=page, dataset=dataset, species=species_data,
        habitats=habitat_data,
    )


def report_17(dataset_id, page):
    dataset = models.Dataset.query.get_or_404(dataset_id)
    species, habitat = get_report_data(dataset)

    species_data = {
        'range': {
            'real': species.filter(
                (models.DataSpeciesRegion.range_reasons_for_change_a == 1)
            ).count(),
            'newinfo': species.filter(
                (models.DataSpeciesRegion.range_reasons_for_change_b == 1)
            ).count(),
            'other': species.filter(
                (models.DataSpeciesRegion.range_reasons_for_change_c == 1)
            ).count(),
        },
        'population': {
            'real': species.filter(
                (models.DataSpeciesRegion.population_reasons_for_change_a == 1)
            ).count(),
            'newinfo': species.filter(
                (models.DataSpeciesRegion.population_reasons_for_change_b == 1)
            ).count(),
            'other': species.filter(
                (models.DataSpeciesRegion.population_reasons_for_change_c == 1)
            ).count(),
        },
        'habitat': {
            'real': species.filter(
                (models.DataSpeciesRegion.habitat_reasons_for_change_a == 1)
            ).count(),
            'newinfo': species.filter(
                (models.DataSpeciesRegion.habitat_reasons_for_change_b == 1)
            ).count(),
            'other': species.filter(
                (models.DataSpeciesRegion.habitat_reasons_for_change_c == 1)
            ).count(),
        },
    }
    species_count = species.count() or 1
    for param, data in species_data.iteritems():
        for reason, value in data.iteritems():
            species_data[param][reason] = value * 100.0 / species_count

    habitat_data = {
        'range': {
            'real': habitat.filter(
                (models.DataHabitattypeRegion.range_reasons_for_change_a == 1)
            ).count(),
            'newinfo': habitat.filter(
                (models.DataHabitattypeRegion.range_reasons_for_change_b == 1)
            ).count(),
            'other': habitat.filter(
                (models.DataHabitattypeRegion.range_reasons_for_change_c == 1)
            ).count(),
        },
        'area': {
            'real': habitat.filter(
                (models.DataHabitattypeRegion.area_reasons_for_change_a == 1)
            ).count(),
            'newinfo': habitat.filter(
                (models.DataHabitattypeRegion.area_reasons_for_change_b == 1)
            ).count(),
            'other': habitat.filter(
                (models.DataHabitattypeRegion.area_reasons_for_change_c == 1)
            ).count(),
        },
    }
    habitat_count = habitat.count() or 1
    for param, data in habitat_data.iteritems():
        for reason, value in data.iteritems():
            habitat_data[param][reason] = value * 100.0 / habitat_count

    return render_template(
        'aggregation/reports/17.html',
        page=page, dataset=dataset, species=species_data, habitat=habitat_data,
    )


def report_16(dataset_id, page):
    dataset = models.Dataset.query.get_or_404(dataset_id)

    datasets = set(
        list(models.Dataset.query.filter_by(status=STATUS_CLOSED)) + [dataset]
    )
    for ds in datasets:
        species, habitat = get_report_data(ds)

        species_count = species.count() or 1
        habitat_count = habitat.count() or 1

        groups = list(models.LuGrupSpecie.query.all())
        CONC = CONCLUSIONS.keys() + [None]
        species_groups = {group: {key: 0 for key in CONC} for group in groups}
        for spec in species:
            group = spec.species and spec.species.lu and spec.species.lu.group
            conc = spec.conclusion_assessment or None
            if group:
                species_groups[group][conc] += 1
        for group, species in species_groups.iteritems():
            for key, value in species.iteritems():
                species[key] = value * 100.0 / species_count

        habitat_groups = {'Habitate': {key: 0 for key in CONC}}
        for hab in habitat:
            conc = hab.conclusion_assessment or None
            habitat_groups['Habitate'][conc] += 1
        for group, habitat in habitat_groups.iteritems():
            for key, value in habitat.iteritems():
                habitat[key] = value * 100.0 / habitat_count

        species_groups = list(species_groups.iteritems())
        species_groups.sort(key=lambda a: a[0].description)
        ds.species_groups = species_groups
        ds.habitat_groups = habitat_groups

    return render_template(
        'aggregation/reports/16.html',
        page=page, dataset=dataset, datasets=datasets,
    )


def report_statistics(dataset_id, page):
    dataset = models.Dataset.query.get_or_404(dataset_id)

    species, habitats = get_report_data(dataset)

    def reports(ds):
        data = {'species': {}, 'habitat': {}}
        for k, v in CONCLUSIONS.iteritems():
            data['species'][k] = (
                species
                .filter(
                    models.DataSpeciesRegion.conclusion_assessment == k).count()
            )
            data['habitat'][k] = (
                habitats
                .filter(
                    models.DataHabitattypeRegion.conclusion_assessment == k).count()
            )
        data['species']['total'] = (
            species.count()
        )
        data['habitat']['total'] = (
            habitats.count()
        )

        species_cl, habitats_cl = get_checklist_data(ds)
        data['species']['total_species'] = (
            species_cl
            .with_entities(
                func.count(models.DataSpeciesCheckList.species_name))
            .group_by(models.DataSpeciesCheckList.species_name).count())

        data['species']['total_reports'] = (
            species_cl.count()
        )

        data['habitat']['total_habitats'] = (
            habitats_cl
            .with_entities(func.count(models.DataHabitatsCheckList.valid_name))
            .group_by(models.DataHabitatsCheckList.valid_name).count())

        data['habitat']['total_reports'] = (
            habitats_cl.count()
        )

        data['missing'] = aggregation_missing_data_report(dataset_id)
        return data

    dataset.reports = reports(dataset)
    return render_template(
        'aggregation/reports/statistics.html',
        page=page, dataset=dataset,
    )


pages_to_views = {
    'raport1': report_missing,
    'raport2': report_validation,
    'raport3': report_bioreg_annex,
    'raport4': report_conservation_status,
    'raport5': report_bioreg_global,
    'raport6': report_13,
    'raport7': report_14,
    'raport8': report_15,
    'raport9': report_16,
    'raport10': report_17,
    'raport11': report_pressures1,
    'raport12': report_measures_high_importance,
    'raport13': report_measures_effects,
    'raport14': report_quality,
    'raport15': report_statistics,
}


@aggregation.route('/raport/<int:dataset_id>/<page>')
@require(perm_view_reports())
@download()
def report_view(dataset_id, page):
    if page not in pages_to_views:
        abort(404)
    return pages_to_views[page](dataset_id, page)


@aggregation.route('/<int:dataset_id>/export_excel/', methods=('GET', 'POST'))
@aggregation.route('/<int:dataset_id>/export_excel/<page>',
                   methods=('GET', 'POST'))
def export_service(dataset_id, page=None):
    report_name = get_report_name(page, dataset_id)
    html = request.form.get('html') or ''
    return get_excel_document(html, report_name)
